#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试用例 Excel 生成器 - 支持自定义模板和记忆系统
依赖：pip install openpyxl
"""

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("错误：请先安装 openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

DEFAULT_COLUMNS = ['用例编号', '模块名称', '用例标题', '优先级', '关联需求ID',
                   '设计方法', '前置条件', '测试步骤', '预期结果', '实际结果',
                   '是否通过', '回归类型', '备注']
DEFAULT_WIDTHS = [15, 12, 28, 8, 14, 10, 22, 45, 35, 25, 12, 10, 18]

PRIORITY_COLORS = {
    'P0': 'FF0000',  # Red
    'P1': 'FF6600',  # Orange
    'P2': 'FFCC00',  # Yellow
    'P3': '99CC00',  # Green
}

REGRESSION_TYPES = ['冒烟', '核心', '全量']
DESIGN_METHODS = ['EP', 'BVA', 'ST', 'EG', 'EP+BVA']

# 字段名称标准化映射
FIELD_MAP = {
    'id': '用例编号', '编号': '用例编号', 'case id': '用例编号',
    'tc_id': '用例编号', '测试编号': '用例编号',
    '模块': '模块名称', 'module': '模块名称', '功能模块': '模块名称',
    '标题': '用例标题', 'title': '用例标题', '用例名称': '用例标题', '测试点': '用例标题',
    'priority': '优先级', '级别': '优先级', '用例级别': '优先级',
    'req_id': '关联需求ID', '需求id': '关联需求ID', 'requirement': '关联需求ID',
    '需求编号': '关联需求ID', '关联需求': '关联需求ID',
    'method': '设计方法', 'design_method': '设计方法', '测试方法': '设计方法',
    '前提条件': '前置条件', 'precondition': '前置条件', '测试前提': '前置条件',
    '步骤': '测试步骤', 'steps': '测试步骤', '操作步骤': '测试步骤', '执行步骤': '测试步骤',
    '期望结果': '预期结果', 'expected': '预期结果', 'expected result': '预期结果',
    '执行结果': '实际结果', 'actual': '实际结果', 'actual result': '实际结果',
    '结果': '是否通过', 'status': '是否通过', 'pass/fail': '是否通过', '状态': '是否通过',
    'regression': '回归类型', 'regression_type': '回归类型', '回归': '回归类型',
    '说明': '备注', 'remark': '备注', 'notes': '备注', '其他': '备注',
}


def learn_template(template_path: str) -> dict:
    """学习用户模板结构，返回 schema"""
    wb = load_workbook(template_path)
    ws = wb.active

    columns = []
    column_widths = []

    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value:
            col_name = str(cell.value).strip()
            columns.append(col_name)
            # 获取列宽
            col_letter = chr(64 + col_idx)
            width = ws.column_dimensions[col_letter].width or 12
            column_widths.append(width)

    # 检测 ID 格式（如果有示例行）
    id_format = None
    if ws.max_row >= 2:
        first_id = ws.cell(row=2, column=1).value
        if first_id and isinstance(first_id, str):
            # 尝试识别格式
            if '_' in first_id:
                id_format = "TC_{MODULE}_{SEQ:03d}"
            elif '-' in first_id:
                id_format = "{MODULE}-{SEQ:03d}"

    schema = {
        "source": template_path,
        "columns": columns,
        "widths": column_widths,
        "id_format": id_format,
        "learned_at": str(Path(template_path).stat().st_mtime)
    }

    wb.close()
    return schema


def create_excel(output: str, data: list, template: str = None, schema: dict = None):
    """生成测试用例 Excel"""

    if template and Path(template).exists():
        # 基于用户模板
        wb = load_workbook(template)
        ws = wb.active
        columns = [cell.value for cell in ws[1] if cell.value]
        start_row = 2

        # 清除示例数据
        for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
            for cell in row:
                cell.value = None
    else:
        # 使用默认格式或 schema
        columns = schema.get("columns", DEFAULT_COLUMNS) if schema else DEFAULT_COLUMNS
        widths = schema.get("widths", DEFAULT_WIDTHS) if schema else DEFAULT_WIDTHS
        start_row = 2

        wb = Workbook()
        ws = wb.active
        ws.title = "测试用例"

        # 写入表头
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # 设置列宽
        for i, width in enumerate(widths):
            if i < len(columns):
                ws.column_dimensions[chr(65 + i)].width = width

    # 建立列索引
    col_index = {}
    for i, col_name in enumerate(columns):
        normalized = col_name.lower().strip()
        standard_name = FIELD_MAP.get(normalized, col_name)
        col_index[standard_name] = i + 1
        col_index[col_name] = i + 1

    # 写入数据
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for row_idx, case in enumerate(data, start_row):
        for key, value in case.items():
            col_num = col_index.get(key)
            if col_num:
                cell = ws.cell(row=row_idx, column=col_num, value=value)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                cell.border = border

    ws.freeze_panes = 'A2'
    wb.save(output)
    print(f"已生成: {output}")


def add_data_validation(ws, start_row: int, end_row: int, col_index: dict):
    """添加数据验证（下拉列表）"""
    # 优先级下拉
    if '优先级' in col_index:
        priority_col = col_index['优先级']
        priority_dv = DataValidation(type="list", formula1='"P0,P1,P2,P3"', allow_blank=True)
        priority_dv.error = '请选择有效的优先级'
        priority_dv.errorTitle = '无效输入'
        ws.add_data_validation(priority_dv)
        for row in range(start_row, end_row + 1):
            priority_dv.add(ws.cell(row=row, column=priority_col))

    # 回归类型下拉
    if '回归类型' in col_index:
        regression_col = col_index['回归类型']
        regression_dv = DataValidation(type="list", formula1='"冒烟,核心,全量"', allow_blank=True)
        regression_dv.error = '请选择有效的回归类型'
        regression_dv.errorTitle = '无效输入'
        ws.add_data_validation(regression_dv)
        for row in range(start_row, end_row + 1):
            regression_dv.add(ws.cell(row=row, column=regression_col))

    # 是否通过下拉
    if '是否通过' in col_index:
        pass_col = col_index['是否通过']
        pass_dv = DataValidation(type="list", formula1='"通过,未通过,阻塞,未执行"', allow_blank=True)
        pass_dv.error = '请选择有效的状态'
        pass_dv.errorTitle = '无效输入'
        ws.add_data_validation(pass_dv)
        for row in range(start_row, end_row + 1):
            pass_dv.add(ws.cell(row=row, column=pass_col))


def apply_priority_colors(ws, start_row: int, end_row: int, col_index: dict):
    """根据优先级设置单元格颜色"""
    if '优先级' not in col_index:
        return

    priority_col = col_index['优先级']
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=priority_col)
        priority = str(cell.value).upper() if cell.value else ''
        if priority in PRIORITY_COLORS:
            cell.fill = PatternFill(start_color=PRIORITY_COLORS[priority],
                                    end_color=PRIORITY_COLORS[priority],
                                    fill_type="solid")
            if priority in ['P0', 'P1']:
                cell.font = Font(bold=True, color="FFFFFF")


def create_traceability_sheet(wb, data: list, requirements: list = None):
    """创建需求追溯矩阵 Sheet"""
    ws = wb.create_sheet(title="需求追溯矩阵")

    # 构建需求-用例映射
    req_case_map = {}
    for case in data:
        req_id = case.get('关联需求ID') or case.get('req_id') or ''
        if req_id:
            req_ids = [r.strip() for r in str(req_id).split(',')]
            for rid in req_ids:
                if rid:
                    if rid not in req_case_map:
                        req_case_map[rid] = []
                    case_id = case.get('用例编号') or case.get('id') or ''
                    if case_id:
                        req_case_map[rid].append(case_id)

    # 如果提供了需求列表，合并
    if requirements:
        for req in requirements:
            req_id = req.get('id') or req.get('需求ID') or ''
            if req_id and req_id not in req_case_map:
                req_case_map[req_id] = []

    # 写入表头
    headers = ['需求ID', '需求名称', '所属模块', '关联用例', '用例数量', '覆盖状态']
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # 写入数据
    row_idx = 2
    covered_count = 0
    total_count = len(req_case_map)

    for req_id, case_ids in sorted(req_case_map.items()):
        case_count = len(case_ids)
        covered = case_count > 0
        if covered:
            covered_count += 1

        ws.cell(row=row_idx, column=1, value=req_id).border = border
        ws.cell(row=row_idx, column=2, value='').border = border  # 需求名称需从 requirements 获取
        ws.cell(row=row_idx, column=3, value='').border = border  # 所属模块
        ws.cell(row=row_idx, column=4, value=', '.join(case_ids)).border = border
        ws.cell(row=row_idx, column=5, value=case_count).border = border

        status_cell = ws.cell(row=row_idx, column=6)
        status_cell.value = '✅ 已覆盖' if covered else '❌ 未覆盖'
        status_cell.border = border
        if not covered:
            status_cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

        row_idx += 1

    # 设置列宽
    widths = [14, 25, 15, 40, 10, 12]
    for i, width in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = width

    ws.freeze_panes = 'A2'
    return covered_count, total_count


def create_coverage_stats_sheet(wb, data: list, covered_count: int, total_req_count: int):
    """创建覆盖率统计 Sheet"""
    ws = wb.create_sheet(title="覆盖率统计")

    # 计算统计数据
    total_cases = len(data)
    priority_counts = {'P0': 0, 'P1': 0, 'P2': 0, 'P3': 0}
    regression_counts = {'冒烟': 0, '核心': 0, '全量': 0}
    orphan_count = 0

    for case in data:
        priority = str(case.get('优先级') or case.get('priority') or '').upper()
        if priority in priority_counts:
            priority_counts[priority] += 1

        regression = case.get('回归类型') or case.get('regression') or ''
        if regression in regression_counts:
            regression_counts[regression] += 1

        req_id = case.get('关联需求ID') or case.get('req_id') or ''
        if not req_id:
            orphan_count += 1

    coverage_rate = (covered_count / total_req_count * 100) if total_req_count > 0 else 0
    coverage_depth = (total_cases / total_req_count) if total_req_count > 0 else 0

    # 写入统计数据
    stats = [
        ('统计项', '数值'),
        ('总需求数', total_req_count),
        ('已覆盖需求数', covered_count),
        ('未覆盖需求数', total_req_count - covered_count),
        ('需求覆盖率', f'{coverage_rate:.1f}%'),
        ('', ''),
        ('总用例数', total_cases),
        ('覆盖深度', f'{coverage_depth:.2f} 用例/需求'),
        ('孤儿用例数', orphan_count),
        ('', ''),
        ('P0 用例数', priority_counts['P0']),
        ('P1 用例数', priority_counts['P1']),
        ('P2 用例数', priority_counts['P2']),
        ('P3 用例数', priority_counts['P3']),
        ('', ''),
        ('冒烟测试用例', regression_counts['冒烟']),
        ('核心回归用例', regression_counts['核心']),
        ('全量回归用例', regression_counts['全量']),
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for row_idx, (label, value) in enumerate(stats, 1):
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        value_cell = ws.cell(row=row_idx, column=2, value=value)

        if row_idx == 1:
            label_cell.font = header_font
            label_cell.fill = header_fill
            value_cell.font = header_font
            value_cell.fill = header_fill

        label_cell.border = border
        value_cell.border = border
        label_cell.alignment = Alignment(horizontal='left', vertical='center')
        value_cell.alignment = Alignment(horizontal='center', vertical='center')

    # 设置列宽
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 20


def main():
    parser = argparse.ArgumentParser(description='生成测试用例 Excel')
    parser.add_argument('-o', '--output', required=True, help='输出文件路径')
    parser.add_argument('-d', '--data', required=True, help='JSON 格式测试用例数据')
    parser.add_argument('-t', '--template', help='模板文件路径')
    parser.add_argument('-s', '--schema', help='从 .memory 读取的 schema JSON')
    parser.add_argument('--learn', help='学习模板结构并输出 schema')
    parser.add_argument('--traceability', action='store_true',
                        help='生成需求追溯矩阵和覆盖率统计 Sheet')
    parser.add_argument('-r', '--requirements', help='需求列表 JSON（用于追溯矩阵）')
    args = parser.parse_args()

    try:
        if args.learn:
            # 学习模式
            schema = learn_template(args.learn)
            print(json.dumps(schema, ensure_ascii=False, indent=2))
            return

        cases = json.loads(args.data)
        schema = json.loads(args.schema) if args.schema else None
        requirements = json.loads(args.requirements) if args.requirements else None

        # 创建 Excel（不保存，先添加其他 Sheet）
        if args.template and Path(args.template).exists():
            wb = load_workbook(args.template)
            ws = wb.active
            columns = [cell.value for cell in ws[1] if cell.value]
            start_row = 2

            for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
                for cell in row:
                    cell.value = None
        else:
            columns = schema.get("columns", DEFAULT_COLUMNS) if schema else DEFAULT_COLUMNS
            widths = schema.get("widths", DEFAULT_WIDTHS) if schema else DEFAULT_WIDTHS
            start_row = 2

            wb = Workbook()
            ws = wb.active
            ws.title = "测试用例"

            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            for col, name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col, value=name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            for i, width in enumerate(widths):
                if i < len(columns):
                    ws.column_dimensions[chr(65 + i)].width = width

        # 建立列索引
        col_index = {}
        for i, col_name in enumerate(columns):
            normalized = col_name.lower().strip()
            standard_name = FIELD_MAP.get(normalized, col_name)
            col_index[standard_name] = i + 1
            col_index[col_name] = i + 1

        # 写入数据
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for row_idx, case in enumerate(cases, start_row):
            for key, value in case.items():
                col_num = col_index.get(key)
                if col_num:
                    cell = ws.cell(row=row_idx, column=col_num, value=value)
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                    cell.border = border

        end_row = start_row + len(cases) - 1 if cases else start_row

        # 添加数据验证和优先级颜色
        add_data_validation(ws, start_row, end_row, col_index)
        apply_priority_colors(ws, start_row, end_row, col_index)

        ws.freeze_panes = 'A2'

        # 生成追溯矩阵和覆盖率统计（如果启用）
        if args.traceability:
            covered_count, total_req_count = create_traceability_sheet(wb, cases, requirements)
            create_coverage_stats_sheet(wb, cases, covered_count, total_req_count)

        wb.save(args.output)
        print(f"已生成: {args.output}")

    except json.JSONDecodeError as e:
        print(f"JSON 解析错误: {e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"生成失败: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
